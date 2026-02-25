"""Excel summary output formatter"""

import logging
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .base_formatter import BaseFormatter
from ..models.workbook import WorkbookModel
from ..utils.logging_utils import get_logger

logger = get_logger(__name__)


class ExcelFormatter(BaseFormatter):
    """Formats workbook analysis as an Excel summary file"""

    def format(self, workbook_model: WorkbookModel, output_path: str, verbose: bool = False):
        """
        Generate Excel summary output.

        Args:
            workbook_model: WorkbookModel to format
            output_path: Output file path
            verbose: If True, log progress

        Returns:
            Output file path
        """
        if verbose:
            logger.info(f"Generating Excel summary: {output_path}")

        # Create new workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create summary sheet
        self._create_summary_sheet(wb, workbook_model)

        # Create a sheet for each worksheet analysis
        for ws_model in workbook_model.worksheets:
            self._create_worksheet_analysis_sheet(wb, ws_model)

        # Save workbook
        output_path = Path(output_path)
        wb.save(output_path)

        if verbose:
            logger.info(f"Excel summary written to: {output_path}")

        return output_path

    def _create_summary_sheet(self, wb, workbook_model):
        """Create summary overview sheet"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "Excel Workbook Analysis Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        # File information
        row = 3
        ws[f'A{row}'] = "File Path:"
        ws[f'B{row}'] = workbook_model.file_path
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "File Format:"
        ws[f'B{row}'] = workbook_model.file_format.upper()
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Total Worksheets:"
        ws[f'B{row}'] = len(workbook_model.worksheets)
        ws[f'A{row}'].font = Font(bold=True)

        # Properties
        row += 2
        ws[f'A{row}'] = "Workbook Properties"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        row += 1
        props = workbook_model.properties
        if props.creator:
            ws[f'A{row}'] = "Creator:"
            ws[f'B{row}'] = props.creator
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        if props.created:
            ws[f'A{row}'] = "Created:"
            ws[f'B{row}'] = str(props.created)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        if props.modified:
            ws[f'A{row}'] = "Modified:"
            ws[f'B{row}'] = str(props.modified)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Worksheet summary
        row += 2
        ws[f'A{row}'] = "Worksheet Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        row += 1
        ws[f'A{row}'] = "Name"
        ws[f'B{row}'] = "Cells"
        ws[f'C{row}'] = "Charts"
        ws[f'D{row}'] = "Images"
        ws[f'E{row}'] = "Merged Cells"
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[f'{col}{row}'].font = Font(color="FFFFFF", bold=True)

        for ws_model in workbook_model.worksheets:
            row += 1
            ws[f'A{row}'] = ws_model.name
            ws[f'B{row}'] = len(ws_model.cells)
            ws[f'C{row}'] = len(ws_model.charts)
            ws[f'D{row}'] = len(ws_model.images)
            ws[f'E{row}'] = len(ws_model.merged_cells)

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

    def _create_worksheet_analysis_sheet(self, wb, ws_model):
        """Create analysis sheet for a worksheet"""
        # Sanitize sheet name (max 31 chars, no special chars)
        sheet_name = ws_model.name[:28] + "..." if len(ws_model.name) > 31 else ws_model.name
        sheet_name = "".join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in sheet_name)

        ws = wb.create_sheet(sheet_name)

        # Title
        ws['A1'] = f"Analysis: {ws_model.name}"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        # Basic info
        ws[f'A{row}'] = "Total Cells:"
        ws[f'B{row}'] = len(ws_model.cells)
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Merged Cells:"
        ws[f'B{row}'] = len(ws_model.merged_cells)
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Charts:"
        ws[f'B{row}'] = len(ws_model.charts)
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Images:"
        ws[f'B{row}'] = len(ws_model.images)
        ws[f'A{row}'].font = Font(bold=True)

        # Cell data table (first 100 cells)
        if ws_model.cells:
            row += 2
            ws[f'A{row}'] = "Cell Data (First 100 cells)"
            ws[f'A{row}'].font = Font(size=12, bold=True)

            row += 1
            headers = ['Cell', 'Type', 'Value', 'Formula', 'Format']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            for cell_model in ws_model.cells[:100]:
                row += 1
                ws.cell(row=row, column=1, value=cell_model.coordinate)
                ws.cell(row=row, column=2, value=cell_model.data_type)
                ws.cell(row=row, column=3, value=str(cell_model.value)[:100])
                ws.cell(row=row, column=4, value=cell_model.formula if cell_model.formula else "")
                ws.cell(row=row, column=5, value=cell_model.number_format)

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
