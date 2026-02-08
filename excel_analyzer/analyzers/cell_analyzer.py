"""Analyzer for cell data and formulas"""

import logging
from typing import List
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from ..models.cell import CellModel
from ..utils.logging_utils import get_logger

logger = get_logger(__name__)


class CellAnalyzer:
    """Extracts cell data, formulas, and basic properties"""

    def extract_cells(self, worksheet: Worksheet, verbose: bool = False) -> List[CellModel]:
        """
        Extract all cells from a worksheet.

        Args:
            worksheet: openpyxl Worksheet object
            verbose: If True, log detailed progress

        Returns:
            List of CellModel objects
        """
        cells = []

        if verbose:
            logger.info(f"Extracting cells from worksheet: {worksheet.title}")

        # Get the used range
        if worksheet.max_row == 0 or worksheet.max_column == 0:
            if verbose:
                logger.info("Worksheet is empty")
            return cells

        # Import here to avoid circular dependency
        from .format_analyzer import FormatAnalyzer
        format_analyzer = FormatAnalyzer()

        # Iterate through all cells in the used range
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column
        ):
            for cell in row:
                # Skip completely empty cells (no value, no formatting)
                if cell.value is None and not self._has_formatting(cell):
                    continue

                # Determine if cell is merged
                is_merged = False
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        break

                # Get column letter
                column_letter = get_column_letter(cell.column)

                # Extract formula if present
                formula = None
                calculated_value = None
                if cell.data_type == 'f':
                    formula = cell.value
                    calculated_value = cell._value  # Cached calculated value

                # Get hyperlink
                hyperlink = None
                if cell.hyperlink:
                    hyperlink = cell.hyperlink.target

                # Get comment
                comment = None
                if cell.comment:
                    comment = cell.comment.text

                # Extract formatting
                formatting = format_analyzer.extract_cell_formatting(cell)

                # Create cell model
                cell_model = CellModel(
                    coordinate=cell.coordinate,
                    row=cell.row,
                    column=cell.column,
                    column_letter=column_letter,
                    value=cell.value if cell.data_type != 'f' else calculated_value,
                    data_type=cell.data_type,
                    number_format=cell.number_format,
                    formula=formula,
                    calculated_value=calculated_value,
                    is_merged=is_merged,
                    formatting=formatting,
                    hyperlink=hyperlink,
                    comment=comment,
                )

                cells.append(cell_model)

        if verbose:
            logger.info(f"Extracted {len(cells)} cells")

        return cells

    def _has_formatting(self, cell) -> bool:
        """Check if cell has any formatting applied"""
        # Check if cell has non-default font
        if cell.font and (
            cell.font.bold or
            cell.font.italic or
            cell.font.color and hasattr(cell.font.color, 'rgb')
        ):
            return True

        # Check if cell has fill
        if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
            return True

        # Check if cell has borders
        if cell.border and any([
            cell.border.left and cell.border.left.style,
            cell.border.right and cell.border.right.style,
            cell.border.top and cell.border.top.style,
            cell.border.bottom and cell.border.bottom.style,
        ]):
            return True

        return False
