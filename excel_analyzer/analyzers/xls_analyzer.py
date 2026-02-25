"""XLS (legacy) file analyzer"""

import logging
import xlrd
from xlrd.formatting import Format
from tqdm import tqdm
from .base_analyzer import BaseAnalyzer
from ..models.workbook import WorkbookModel, WorkbookPropertiesModel
from ..models.worksheet import WorksheetModel, ColumnDimensionModel, RowDimensionModel
from ..models.cell import CellModel
from ..models.formatting import (
    CellFormattingModel,
    FontModel,
    FillModel,
    BorderModel,
    BorderSideModel,
    AlignmentModel,
    ProtectionModel,
    ColorModel,
)
from ..utils.logging_utils import get_logger

logger = get_logger(__name__)


class XLSAnalyzer(BaseAnalyzer):
    """Analyzer for .xls files using xlrd"""

    # XLS has limited color support - 64 color palette
    EXCEL_COLOR_MAP = {
        0: "#000000",  # Black
        1: "#FFFFFF",  # White
        2: "#FF0000",  # Red
        3: "#00FF00",  # Green
        4: "#0000FF",  # Blue
        5: "#FFFF00",  # Yellow
        6: "#FF00FF",  # Magenta
        7: "#00FFFF",  # Cyan
        # Add more as needed...
    }

    def __init__(self):
        self.workbook = None
        self.formatting_info = True

    def analyze(self, file_path: str, verbose: bool = False) -> WorkbookModel:
        """
        Analyze an XLS file.

        Args:
            file_path: Path to .xls file
            verbose: If True, show detailed progress

        Returns:
            WorkbookModel with extracted data (note: limited compared to XLSX)
        """
        if verbose:
            logger.info(f"Opening XLS workbook: {file_path}")
            logger.warning("XLS format has limited support for charts, images, and conditional formatting")

        # Open workbook with formatting info
        self.workbook = xlrd.open_workbook(file_path, formatting_info=True)

        if verbose:
            logger.info(f"Workbook loaded with {self.workbook.nsheets} worksheets")

        # Extract properties
        properties = self._extract_properties()

        # Create workbook model
        workbook_model = WorkbookModel(
            file_path=file_path,
            file_format='xls',
            properties=properties,
            worksheets=[],
        )

        # Analyze each worksheet
        sheets = range(self.workbook.nsheets)
        if verbose:
            sheets = tqdm(sheets, desc="Analyzing worksheets", unit="sheet")

        for sheet_idx in sheets:
            worksheet = self.workbook.sheet_by_index(sheet_idx)

            if verbose and not isinstance(sheets, tqdm):
                logger.info(f"Analyzing worksheet: {worksheet.name}")

            worksheet_model = self._analyze_worksheet(worksheet, sheet_idx, verbose)
            workbook_model.worksheets.append(worksheet_model)

        if verbose:
            logger.info("Analysis complete")

        return workbook_model

    def _extract_properties(self) -> WorkbookPropertiesModel:
        """Extract workbook properties (limited in XLS)"""
        # XLS has very limited metadata support
        return WorkbookPropertiesModel(
            title=None,
            creator=None,
            created=None,
        )

    def _analyze_worksheet(self, sheet, index: int, verbose: bool) -> WorksheetModel:
        """Analyze a single worksheet"""

        cells = []
        row_dimensions = {}
        column_dimensions = {}

        # Extract cells
        for row_idx in range(sheet.nrows):
            row = sheet.row(row_idx)

            # Get row height if available
            if hasattr(sheet, 'rowinfo_map') and row_idx in sheet.rowinfo_map:
                row_info = sheet.rowinfo_map[row_idx]
                row_dimensions[row_idx + 1] = RowDimensionModel(
                    row=row_idx + 1,
                    height=row_info.height / 20.0,  # Convert twips to points
                    hidden=row_info.hidden if hasattr(row_info, 'hidden') else False,
                )

            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)

                # Skip empty cells
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    continue

                # Get cell value
                value = cell.value
                data_type = self._get_data_type(cell.ctype)

                # Get cell format
                xf_index = sheet.cell_xf_index(row_idx, col_idx) if hasattr(sheet, 'cell_xf_index') else None
                formatting = self._extract_formatting(xf_index) if xf_index is not None else None

                # Get number format
                number_format = "General"
                if xf_index is not None and xf_index < len(self.workbook.format_map):
                    fmt = self.workbook.format_map.get(xf_index)
                    if fmt:
                        number_format = fmt.format_str

                # Create cell coordinate
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(col_idx + 1)
                coordinate = f"{column_letter}{row_idx + 1}"

                cell_model = CellModel(
                    coordinate=coordinate,
                    row=row_idx + 1,
                    column=col_idx + 1,
                    column_letter=column_letter,
                    value=value,
                    data_type=data_type,
                    number_format=number_format,
                    formatting=formatting,
                )

                cells.append(cell_model)

        # Extract column widths
        if hasattr(sheet, 'colinfo_map'):
            for col_idx, col_info in sheet.colinfo_map.items():
                from openpyxl.utils import get_column_letter
                column_letter = get_column_letter(col_idx + 1)
                column_dimensions[column_letter] = ColumnDimensionModel(
                    column=column_letter,
                    width=col_info.width / 256.0,  # Convert to Excel width units
                    hidden=col_info.hidden if hasattr(col_info, 'hidden') else False,
                )

        # Extract merged cells
        merged_cells = []
        if hasattr(sheet, 'merged_cells'):
            for rlo, rhi, clo, chi in sheet.merged_cells:
                from openpyxl.utils import get_column_letter
                start_cell = f"{get_column_letter(clo + 1)}{rlo + 1}"
                end_cell = f"{get_column_letter(chi)}{rhi}"
                merged_cells.append(f"{start_cell}:{end_cell}")

        return WorksheetModel(
            name=sheet.name,
            index=index,
            cells=cells,
            merged_cells=merged_cells,
            column_dimensions=column_dimensions,
            row_dimensions=row_dimensions,
            data_validations=[],  # Not supported in xlrd
            conditional_formatting=[],  # Not supported in xlrd
            charts=[],  # Very limited in XLS
            images=[],  # Very limited in XLS
        )

    def _get_data_type(self, ctype: int) -> str:
        """Convert xlrd cell type to our data type"""
        type_map = {
            xlrd.XL_CELL_EMPTY: 'empty',
            xlrd.XL_CELL_TEXT: 's',
            xlrd.XL_CELL_NUMBER: 'n',
            xlrd.XL_CELL_DATE: 'd',
            xlrd.XL_CELL_BOOLEAN: 'b',
            xlrd.XL_CELL_ERROR: 'e',
            xlrd.XL_CELL_BLANK: 'empty',
        }
        return type_map.get(ctype, 's')

    def _extract_formatting(self, xf_index: int) -> CellFormattingModel:
        """Extract cell formatting from XF record"""
        try:
            if xf_index >= len(self.workbook.format_list):
                return None

            xf = self.workbook.format_list[xf_index]
            font = self.workbook.font_list[xf.font_index] if xf.font_index < len(self.workbook.font_list) else None

            # Extract font
            font_model = FontModel()
            if font:
                font_model = FontModel(
                    name=font.name,
                    size=font.height / 20.0,  # Convert twips to points
                    bold=font.bold if hasattr(font, 'bold') else False,
                    italic=font.italic if hasattr(font, 'italic') else False,
                    underline="single" if (hasattr(font, 'underline_type') and font.underline_type) else "none",
                    strike=font.struck_out if hasattr(font, 'struck_out') else False,
                    color=self._get_color(font.colour_index) if hasattr(font, 'colour_index') else None,
                )

            # Extract fill (limited support)
            fill_model = FillModel()
            if hasattr(xf, 'background') and xf.background:
                pattern_type = "solid" if xf.background.pattern_colour_index != 64 else "none"
                fill_model = FillModel(
                    pattern_type=pattern_type,
                    fg_color=self._get_color(xf.background.pattern_colour_index),
                    bg_color=self._get_color(xf.background.background_colour_index),
                )

            # Extract border (limited support)
            border_model = BorderModel()
            if hasattr(xf, 'border'):
                border = xf.border
                border_model = BorderModel(
                    left=BorderSideModel(style=self._get_border_style(border.left_line_style)),
                    right=BorderSideModel(style=self._get_border_style(border.right_line_style)),
                    top=BorderSideModel(style=self._get_border_style(border.top_line_style)),
                    bottom=BorderSideModel(style=self._get_border_style(border.bottom_line_style)),
                )

            # Extract alignment
            alignment_model = AlignmentModel()
            if hasattr(xf, 'alignment'):
                align = xf.alignment
                h_align_map = {0: "general", 1: "left", 2: "center", 3: "right", 4: "fill", 5: "justify"}
                v_align_map = {0: "top", 1: "center", 2: "bottom", 3: "justify"}

                alignment_model = AlignmentModel(
                    horizontal=h_align_map.get(align.hor_align, "general"),
                    vertical=v_align_map.get(align.vert_align, "bottom"),
                    wrap_text=align.text_wrapped if hasattr(align, 'text_wrapped') else False,
                    text_rotation=align.rotation if hasattr(align, 'rotation') else 0,
                )

            # Extract protection
            protection_model = ProtectionModel(
                locked=xf.protection.cell_locked if hasattr(xf, 'protection') else True,
                hidden=xf.protection.formula_hidden if hasattr(xf, 'protection') else False,
            )

            return CellFormattingModel(
                font=font_model,
                fill=fill_model,
                border=border_model,
                alignment=alignment_model,
                protection=protection_model,
            )

        except Exception as e:
            logger.warning(f"Error extracting formatting: {e}")
            return None

    def _get_color(self, color_index: int) -> ColorModel:
        """Convert XLS color index to ColorModel"""
        if color_index is None or color_index == 64 or color_index == 65:
            return None

        # Try to get color from palette or use default mapping
        hex_color = self.EXCEL_COLOR_MAP.get(color_index, "#000000")

        return ColorModel(
            type='indexed',
            value=hex_color,
        )

    def _get_border_style(self, style_index: int) -> str:
        """Convert XLS border style index to style name"""
        style_map = {
            0: "none",
            1: "thin",
            2: "medium",
            3: "dashed",
            4: "dotted",
            5: "thick",
            6: "double",
            7: "hair",
        }
        return style_map.get(style_index, "none")
